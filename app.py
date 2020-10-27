import time
import logging
import sched
import threading
import os
from logging.config import dictConfig
from datetime import datetime
from collections import OrderedDict
from uuid import uuid4

from dateutil.parser import parse as dt_parse, ParserError as DateParserError
from dateutil.relativedelta import relativedelta

import click
from tqdm import tqdm
import gspread
from flask import Flask, request, Response, url_for, abort
from sqlalchemy.sql import and_, func, expression
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from translitua import translit

from viberbot import Api
from viberbot.api.bot_configuration import BotConfiguration
from viberbot.api.messages import TextMessage, RichMediaMessage
from viberbot.api.viber_requests import (
    ViberConversationStartedRequest,
    ViberFailedRequest,
    ViberMessageRequest,
    ViberSubscribedRequest,
    ViberUnsubscribedRequest,
)
from viberbot.api.messages import TextMessage, ContactMessage, PictureMessage, VideoMessage, KeyboardMessage


from exc import InvalidSheet, InvalidRecord
from storage import get_postgres_database
from dicts import REGIONS, PRODUCT_CATEGORIES, HEADERS, SUBSCRIPTION_TYPES
from keyboards import VIBER_MENU_KBD, VIBER_REGIONS_KBD, get_viber_categories_kbd, get_viber_subscribe_kbd
from utils import parse_amount, parse_int


app = Flask(__name__)
app.config.from_object("default_settings")
db = get_postgres_database(app)
procurements = db["procurements"]
subscriptions = db["subscriptions"]
sent_log = db["sent_log"]
dictConfig(app.config["LOGGING"])

viber = Api(
    BotConfiguration(
        name=app.config["BOT_NAME"],
        avatar=f"{app.config['WEBHOOK_URL']}/static/avatar.png",
        auth_token=app.config["BOT_AUTH_TOKEN"],
    )
)


def get_product_stats_since(region, product_name, since):
    ptc = procurements.table.c

    q = db.query(
        expression.select(
            [
                func.count(ptc.total_amount).label("count"),
                func.sum(ptc.total_amount).label("total"),
                func.min(ptc.price).label("min"),
                func.avg(ptc.price).label("avg"),
                func.max(ptc.price).label("max"),
            ],
            whereclause=and_(ptc.product_name == product_name, ptc.region == region, ptc.signature_date >= since),
        )
    )

    for r in q:
        if r["count"] > 0:
            return r


def get_product_stats(region, product_name):
    now = datetime.now(app.config["TIMEZONE"])

    periods = (
        ("За останню добу", relativedelta(days=-1)),
        ("За останній тиждень", relativedelta(days=-7)),
        ("За останній місяць", relativedelta(months=-1)),
        ("За весь час", relativedelta(years=-100)),
    )

    res = []
    for label, period in periods:
        r = get_product_stats_since(region, product_name, now + period)

        if r is not None:
            r["since"] = now + period
            res.append((label, r))

    if res:
        return OrderedDict(res)
    else:
        return None


def subscribe_user(user_id, region, product_name, period):
    updated = subscriptions.upsert(
        {
            "user_id": user_id,
            "region": region,
            "product_name": product_name,
            "period": period,
            "uuid": str(uuid4()),
            "dt": datetime.now(app.config["TIMEZONE"]),
        },
        ["user_id", "region", "product_name", "period"],
    )

    if updated == True:
        return False
    return True


def get_active_subscriptions(user_id):
    return list(subscriptions.find(user_id=user_id))


def unsubscribe(user_id, uuid):
    return subscriptions.delete(user_id=user_id, uuid=uuid)


@app.cli.command("sync_spreadsheet")
@click.option("--purge", default=False, is_flag=True)
def sync_spreadsheet(purge):
    inserted_count = 0
    updated_count = 0
    invalid_count = 0
    useful_sheets = 0
    invalid_sheets = 0

    gc = gspread.service_account(os.path.join("keys", app.config["GDRIVE_KEY"]))
    sp = gc.open_by_key(app.config["GDRIVE_SPREADSHEET"])

    if purge:
        procurements.drop()
    for sheet_num, sheet in enumerate(tqdm(sp.worksheets(), desc="Sheets")):
        try:
            for rec in tqdm(sheet.get_all_records(), desc=f"Records in sheet {sheet_num + 1}"):
                refined_rec = {}

                try:
                    for k, v in rec.items():
                        k = k.lower().strip()
                        if isinstance(v, str):
                            v = v.strip()

                        if not k:
                            continue

                        if k not in HEADERS:
                            app.logger.warning(f"Cannot parse header record {k}, aborting current sheet {sheet}")
                            raise InvalidSheet()

                        new_k = HEADERS[k]

                        if new_k == "product_name":
                            if v.lower() not in PRODUCT_CATEGORIES:
                                app.logger.warning(f"Cannot parse product_name {v}, skipping rec {rec}")
                                raise InvalidRecord()

                            refined_rec[new_k] = PRODUCT_CATEGORIES[v.lower()]
                        elif new_k == "region":
                            if v.lower() not in REGIONS:
                                app.logger.warning(f"Cannot parse region {v}, skipping rec {rec}")
                                raise InvalidRecord()

                            refined_rec[new_k] = REGIONS[v.lower()]
                        elif new_k == "product_details":
                            refined_rec[new_k] = v
                            refined_rec["product_hash"] = v.lower().strip()
                        elif new_k in ["price", "total_amount"]:
                            try:
                                refined_rec[new_k] = parse_amount(v)
                            except ValueError:
                                app.logger.warning(f"Cannot parse field {new_k} {v}, skipping rec {rec}")
                                raise InvalidRecord()
                        elif new_k in ["participants"]:
                            try:
                                refined_rec[new_k] = parse_int(v)
                            except ValueError:
                                app.logger.warning(f"Cannot parse number of participants {v}, skipping rec {rec}")
                                raise InvalidRecord()
                        elif new_k in ["signature_date"]:
                            try:
                                refined_rec[new_k] = app.config["TIMEZONE"].localize(dt_parse(v, dayfirst=True))
                            except DateParserError:
                                app.logger.warning(f"Cannot parse date of signature {v}, skipping rec {rec}")
                                raise InvalidRecord()
                        else:
                            refined_rec[new_k] = v

                    update = procurements.upsert(refined_rec, ["contract_id", "product_name", "product_hash"])

                    if update == True:
                        updated_count += 1
                    else:
                        inserted_count += 1
                except InvalidRecord:
                    invalid_count += 1
                    continue

            useful_sheets += 1
        except InvalidSheet as e:
            invalid_sheets += 1

    if purge:
        procurements.create_index(["product_name", "region", "signature_date"])

    app.logger.info(f"Sheets processed: {useful_sheets}, sheets skipped: {invalid_sheets}")
    app.logger.info(
        f"Records added: {inserted_count}, records updated: {updated_count}, records skipped: {invalid_count}"
    )


@app.route("/export/<product_name>/<region>/<since>", methods=["GET"])
def export(product_name, region, since):
    try:
        assert product_name in PRODUCT_CATEGORIES.values()
        assert region in REGIONS.values()
        dt_since = dt_parse(since)
    except (AssertionError, DateParserError):
        abort(403, description="Помилка в параметрах")

    q = procurements.find(
        product_name=product_name, region=region, signature_date={">=": dt_since}, order_by="-signature_date"
    )

    wb = Workbook()
    ws = wb.active
    bold = Font(bold=True)
    ws.title = "Звіт по закупівлях"
    header = [
        "Ідентифікатор договору",
        "Дата підписання",
        "Організатор",
        "Переможець",
        "Сума договору",
        "Кількість учасників",
        "Назва продукту",
        "Характеристика продутку",
        "Ціна за кг",
        "Область та м. київ",
    ]

    for i, h in enumerate(header):
        cell = ws.cell(row=1, column=i + 1, value=h)
        cell.font = bold
        ws.column_dimensions[get_column_letter(i + 1)].width = len(h) + 3

    ws.freeze_panes = "B2"

    for j, r in enumerate(q):
        contract_cell = ws.cell(
            row=j + 2,
            column=1,
            value=r["contract_id"],
        )

        contract_cell.hyperlink = "https://prozorro.gov.ua/tender/{}".format(r["contract_id"][:-3])
        contract_cell.style = "Hyperlink"

        ws.cell(row=j + 2, column=2, value=r["signature_date"])
        ws.cell(row=j + 2, column=3, value=r["buyer"])
        ws.cell(row=j + 2, column=4, value=r["seller"])
        ws.cell(row=j + 2, column=5, value=r["total_amount"])
        ws.cell(row=j + 2, column=6, value=r["participants"])
        ws.cell(row=j + 2, column=7, value=r["product_name"])
        ws.cell(row=j + 2, column=8, value=r["product_details"])
        ws.cell(row=j + 2, column=9, value=r["price"])
        ws.cell(row=j + 2, column=10, value=r["region"])

    return Response(
        save_virtual_workbook(wb),
        headers={
            "Content-Disposition": f"attachment; filename=report_{translit(region).lower()}_{translit(product_name).replace(' ', '_')}.xlsx",
            "Content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )

    return Response("Everything is ok")


@app.route("/", methods=["POST"])
def incoming():
    app.logger.debug(f"received request. post data: {request.get_data()}")

    viber_request = viber.parse_request(request.get_data().decode("utf8"))

    if isinstance(viber_request, ViberMessageRequest):
        message = viber_request.message
        chunks = message.text.split(":")

        if not chunks:
            command = "start"
        else:
            command = chunks[0]

        if command == "start":
            viber.send_messages(
                viber_request.sender.id,
                TextMessage(
                    text="Для того щоб розпочати роботу оберіть внизу область по котрій ви хочете отримувати цінову інформацію",
                    keyboard=VIBER_REGIONS_KBD,
                ),
            )
        elif command == "help":
            viber.send_messages(
                viber_request.sender.id,
                TextMessage(
                    text="Бот дозволяє вам отримувати актуальну інформацію щодо цінових пропозицій на різні категорії товарів а також підписуватися на такі цінові пропозиції",
                    keyboard=VIBER_MENU_KBD,
                ),
            )
        elif command == "subscriptions":
            subs = get_active_subscriptions(viber_request.sender.id)

            if subs:
                carousel = {
                    "ButtonsGroupRows": 5,
                    "ButtonsGroupColumns": 6,
                    "BgColor": "#FFFFFF",
                    "Buttons": [],
                }

                for sub in subs:
                    period = ""
                    for p_readable, p in SUBSCRIPTION_TYPES.items():
                        if sub["period"] == p:
                            period = p_readable
                            break

                    carousel["Buttons"].append(
                        {
                            "ActionBody": f"product_name:{sub['region']}:{sub['product_name']}",
                            "ActionType": "reply",
                            "TextVAlign": "top",
                            "TextHAlign": "left",
                            "Text": f"<b>Категорія</b>: {sub['product_name']}\n<b>Регіон</b>: {sub['region']}\n\n{period}",
                            "Rows": 4,
                            "Columns": 6,
                        }
                    )
                    carousel["Buttons"].append(
                        {
                            "ActionBody": f"unsubscribe:{sub['uuid']}",
                            "ActionType": "reply",
                            "TextVAlign": "middle",
                            "TextHAlign": "middle",
                            "BgColor": "#ff0000",
                            "Text": '<font color="#FFFFFF"><b>Відписатись</b></font>',
                            "Rows": 1,
                            "Columns": 6,
                        }
                    )

                response_message = RichMediaMessage(
                    rich_media=carousel,
                    alt_text="Ваш viber-клієнт дуже застарів, будь ласка, оновить його",
                    min_api_version=2,
                    keyboard=VIBER_MENU_KBD,
                )
            else:
                response_message = TextMessage(text="У вас поки що нема активних підписок", keyboard=VIBER_MENU_KBD)

            viber.send_messages(viber_request.sender.id, response_message)
        elif command == "unsubscribe":
            if unsubscribe(viber_request.sender.id, chunks[1]):
                response_message = TextMessage(text="Ви були успішно відписані", keyboard=VIBER_MENU_KBD)
            else:
                response_message = TextMessage(text="Виникла помилка", keyboard=VIBER_MENU_KBD)
            viber.send_messages(viber_request.sender.id, response_message)
        elif command == "subscribe":
            if (
                len(chunks) < 4
                or chunks[1] not in REGIONS.values()
                or chunks[2] not in PRODUCT_CATEGORIES.values()
                or chunks[3] not in SUBSCRIPTION_TYPES.values()
            ):
                viber.send_messages(
                    viber_request.sender.id,
                    TextMessage(
                        text="Вибачте, не зрозумів, спробуйте почати з початку або подивитися довідку",
                        keyboard=VIBER_MENU_KBD,
                    ),
                )

            if subscribe_user(
                user_id=viber_request.sender.id, region=chunks[1], product_name=chunks[2], period=chunks[3]
            ):
                viber.send_messages(
                    viber_request.sender.id,
                    TextMessage(
                        text=f"Дякую, ви успішно підписані на оновлення по категорії ”{chunks[2]}” в області ”{chunks[1]}”",
                        keyboard=VIBER_MENU_KBD,
                    ),
                )
            else:
                viber.send_messages(
                    viber_request.sender.id,
                    TextMessage(
                        text=f"Ви вже підписані на оновлення по категорії ”{chunks[2]}” в області ”{chunks[1]}”. Ви можете подивитися активні підписки у розділі ”Ваші підписки”",
                        keyboard=VIBER_MENU_KBD,
                    ),
                )
        elif command == "region":
            if len(chunks) > 1 and chunks[1] in REGIONS.values():
                viber.send_messages(
                    viber_request.sender.id,
                    TextMessage(text="Оберіть категорію товару", keyboard=get_viber_categories_kbd(chunks[1])),
                )
            else:
                viber.send_messages(
                    viber_request.sender.id,
                    TextMessage(text="Вибачте, не зрозумів, оберіть, будь ласка, область", keyboard=VIBER_REGIONS_KBD),
                )
        elif command == "product_name":
            if len(chunks) < 3 or chunks[1] not in REGIONS.values() or chunks[2] not in PRODUCT_CATEGORIES.values():
                viber.send_messages(
                    viber_request.sender.id,
                    TextMessage(text="Вибачте, не зрозумів, оберіть, будь ласка, область", keyboard=VIBER_REGIONS_KBD),
                )
            else:
                stats = get_product_stats(chunks[1], chunks[2])

                if stats is None:
                    response_message = TextMessage(
                        text=f"Поки що за вашим запитом ”{chunks[2]}” в області ”{chunks[1]}” нічого не знайдено",
                        keyboard=get_viber_subscribe_kbd(chunks[1], chunks[2]),
                    )

                else:
                    carousel = {
                        "ButtonsGroupRows": 5,
                        "ButtonsGroupColumns": 6,
                        "BgColor": "#FFFFFF",
                        "Buttons": [],
                    }

                    for period, stat in stats.items():
                        report_url = app.config["WEBHOOK_URL"] + url_for(
                            "export", region=chunks[1], product_name=chunks[2], since=stat["since"]
                        )
                        carousel["Buttons"].append(
                            {
                                "ActionBody": report_url,
                                "ActionType": "open-url",
                                "TextVAlign": "top",
                                "TextHAlign": "left",
                                "Text": f"<b>{period}</b>"
                                + f"\n\nВсього закупівель: {stat['count']}\nНа суму: {stat['total']:.2f} грн.\nМінімальна ціна: {stat['min']:.2f} грн."
                                + f"\nМаксимальна ціна: {stat['max']:.2f} грн.\nСередня ціна: {stat['avg']:.2f} грн.\n",
                                "Rows": 4,
                                "Columns": 6,
                            }
                        )
                        carousel["Buttons"].append(
                            {
                                "ActionBody": report_url,
                                "ActionType": "open-url",
                                "TextVAlign": "middle",
                                "TextHAlign": "middle",
                                "BgColor": "#aaaaaa",
                                "Text": "<b>Скачати звіт</b>",
                                "Rows": 1,
                                "Columns": 6,
                            }
                        )

                    response_message = RichMediaMessage(
                        rich_media=carousel,
                        alt_text="Ваш viber-клієнт дуже застарів, будь ласка, оновить його",
                        min_api_version=2,
                    )

                viber.send_messages(
                    viber_request.sender.id,
                    [
                        response_message,
                        TextMessage(
                            text="Ви можете підписатий на оновлення по цій категорії товарів:",
                            keyboard=get_viber_subscribe_kbd(chunks[1], chunks[2]),
                        ),
                    ],
                )

    elif isinstance(viber_request, ViberConversationStartedRequest):
        viber.send_messages(
            viber_request.user.id,
            [
                TextMessage(
                    text="Вітаємо вас в нашому чат-боті Ціновий Вісник! Нажміть Розпочати роботу або скористайтесь Допомогою",
                    keyboard=VIBER_MENU_KBD,
                ),
            ],
        )
    elif isinstance(viber_request, ViberSubscribedRequest):
        viber.send_messages(viber_request.sender.id, [TextMessage(None, None, viber_request.get_event_type())])
    elif isinstance(viber_request, ViberFailedRequest):
        app.logger.warning("client failed receiving message. failure: {viber_request}")

    return Response(status=200)


@app.cli.command("send_subscriptions")
def send_subscriptions():
    now = datetime.now(app.config["TIMEZONE"])

    periods = ["daily"]
    if now.date().weekday == 0:
        periods.append("weekly")
    if now.day == 1:
        periods.append("monthly")

    offsets = {"daily": relativedelta(days=-1), "weekly": relativedelta(days=-7), "monthly": relativedelta(months=-1)}

    first_insert = True

    for period in periods:
        sent_stats = 0
        since = now + offsets[period]
        for sub in subscriptions.find(period=period):
            if sent_log.find_one(subscription_id=sub["id"], dt=now.date()):
                app.logger.info(f"Skipping subscription {sub['id']} as it was already processed today ({now.date()})")
                continue

            try:
                stat = get_product_stats_since(sub["region"], sub["product_name"], since)

                carousel = {
                    "ButtonsGroupRows": 5,
                    "ButtonsGroupColumns": 6,
                    "BgColor": "#FFFFFF",
                    "Buttons": [],
                }

                report_url = url_for(
                    "export", region=sub["region"], product_name=sub["product_name"], since=since
                )
                carousel["Buttons"].append(
                    {
                        "ActionBody": report_url,
                        "ActionType": "open-url",
                        "TextVAlign": "top",
                        "TextHAlign": "left",
                        "Text": f"<b>Ваша підписка на ”{sub['product_name']}” в області ”{sub['region']}”</b>"
                        + f"\n\nВсього закупівель: {stat['count']}\nНа суму: {stat['total']:.2f} грн.\nМінімальна ціна: {stat['min']:.2f} грн."
                        + f"\nМаксимальна ціна: {stat['max']:.2f} грн.\nСередня ціна: {stat['avg']:.2f} грн.\n"
                        + "Звіт за: {}-{}".format(since.strftime(app.config["DT_FORMAT"]), now.strftime(app.config["DT_FORMAT"])),
                        "Rows": 4,
                        "Columns": 6,
                    }
                )
                carousel["Buttons"].append(
                    {
                        "ActionBody": report_url,
                        "ActionType": "open-url",
                        "TextVAlign": "middle",
                        "TextHAlign": "middle",
                        "BgColor": "#aaaaaa",
                        "Text": "<b>Скачати звіт</b>",
                        "Rows": 1,
                        "Columns": 6,
                    }
                )

                viber.send_messages(
                    sub["user_id"],
                    [
                        RichMediaMessage(
                            rich_media=carousel,
                            alt_text="Ваш viber-клієнт дуже застарів, будь ласка, оновить його",
                            min_api_version=2,
                            keyboard=VIBER_MENU_KBD,
                        )
                    ],
                )
                sent_log.insert({"subscription_id": sub["id"], "dt": now.date(), "status": "ok"})
                sent_stats += 1
            except Exception as e:
                sent_log.insert({"subscription_id": sub["id"], "dt": now.date(), "status": "fail"})
                app.logger.error(f"Subscription {sub['id']} raised an error '{e}'")
                raise

            if first_insert:
                sent_log.create_index(["subscription_id", "dt"])

            first_insert = False

        app.logger.info(f"{sent_stats} has been sent successfuly for period {period}")


if __name__ == "__main__":
    def set_webhook(viber):
        app.logger.info("Setting webhook")
        viber.set_webhook(app.config["WEBHOOK_URL"])

    scheduler = sched.scheduler(time.time, time.sleep)
    scheduler.enter(5, 1, set_webhook, (viber,))
    t = threading.Thread(target=scheduler.run)
    t.start()

    app.run(host="0.0.0.0", debug=app.config["DEBUG"])
